"""
Enhanced Excel Extractor - Preserves merged cells structure
Outputs plain text format instead of markdown tables
"""

import os
from typing import Dict, Any, List, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell

class EnhancedExcelExtractor:
    """
    Enhanced Excel extractor that preserves merged cells structure
    and outputs in plain text format
    """
    
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xls']
    
    def extract(self, file_path: str) -> Dict[str, Any]:
        """
        Extract content from Excel file preserving merged cells structure
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        if not self.is_supported_file(file_path):
            raise ValueError(f"Unsupported file format. Supported: {', '.join(self.supported_extensions)}")
        
        result = {
            'filename': os.path.basename(file_path),
            'file_type': 'Excel',
            'content': '',
            'metadata': {},
            'sheets': {}
        }
        
        try:
            # Load workbook
            print(f"Loading Excel workbook...")
            workbook = load_workbook(file_path, data_only=True)
            
            # Process each worksheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                print(f"Processing sheet: {sheet_name}")
                
                sheet_content = self._extract_sheet_content(sheet)
                result['sheets'][sheet_name] = sheet_content
            
            # Combine all sheets content
            result['content'] = self._combine_sheets_content(result['sheets'])
            
            # Extract metadata
            result['metadata'] = {
                'sheets_count': len(workbook.sheetnames),
                'sheet_names': workbook.sheetnames,
                'extraction_method': 'Enhanced Excel Extractor',
                'preserves_merged_cells': True
            }
            
            print(f"Excel processed successfully - {result['metadata']['sheets_count']} sheets")
            
        except Exception as e:
            raise Exception(f"Error processing Excel file: {str(e)}")
        
        return result
    
    def _extract_sheet_content(self, sheet: Worksheet) -> Dict[str, Any]:
        """
        Extract content from a single worksheet preserving structure
        """
        # Get the used range
        min_row, max_row = 1, sheet.max_row
        min_col, max_col = 1, sheet.max_column
        
        # Create a grid to represent the sheet structure
        grid = []
        merged_ranges = list(sheet.merged_cells.ranges)
        
        # Initialize grid with empty values
        for row in range(max_row):
            grid.append([''] * max_col)
        
        # Fill grid with cell values
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                
                if not isinstance(cell, MergedCell):
                    value = cell.value
                    if value is not None:
                        grid[row-1][col-1] = str(value)
        
        # Process merged cells
        merged_info = []
        for merged_range in merged_ranges:
            # Get the top-left cell value
            top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            value = top_left_cell.value
            if value is not None:
                merged_info.append({
                    'range': str(merged_range),
                    'value': str(value),
                    'start_row': merged_range.min_row,
                    'end_row': merged_range.max_row,
                    'start_col': merged_range.min_col,
                    'end_col': merged_range.max_col
                })
        
        return {
            'grid': grid,
            'merged_cells': merged_info,
            'dimensions': {
                'rows': max_row,
                'cols': max_col
            }
        }
    
    def _combine_sheets_content(self, sheets: Dict[str, Any]) -> str:
        """
        Combine all sheets content into plain text format
        """
        content_parts = []
        
        for sheet_name, sheet_data in sheets.items():
            content_parts.append(f"=== SHEET: {sheet_name} ===\n")
            
            # Add merged cells information
            if sheet_data['merged_cells']:
                content_parts.append("--- MERGED CELLS STRUCTURE ---")
                for merged in sheet_data['merged_cells']:
                    content_parts.append(f"Range {merged['range']}: {merged['value']}")
                content_parts.append("")
            
            # Add grid content in structured format
            content_parts.append("--- SHEET CONTENT ---")
            grid = sheet_data['grid']
            
            # Find non-empty rows and columns
            non_empty_rows = []
            for i, row in enumerate(grid):
                if any(cell.strip() for cell in row if cell):
                    non_empty_rows.append(i)
            
            if non_empty_rows:
                # Determine column widths for better formatting
                max_col_width = 20  # Maximum column width
                
                for row_idx in non_empty_rows:
                    row = grid[row_idx]
                    row_content = []
                    
                    for col_idx, cell in enumerate(row):
                        if cell.strip():  # Only show non-empty cells
                            # Truncate long content
                            display_value = cell[:max_col_width] + "..." if len(cell) > max_col_width else cell
                            row_content.append(f"Col{col_idx+1}: {display_value}")
                    
                    if row_content:
                        content_parts.append(f"Row {row_idx+1}: {' | '.join(row_content)}")
            
            content_parts.append("\n" + "="*60 + "\n")
        
        return "\n".join(content_parts)
    
    def to_plain_text(self, extracted_data: Dict[str, Any]) -> str:
        """
        Convert extracted data to structured plain text
        """
        content_parts = []
        
        # Add header information
        content_parts.append("EXCEL DOCUMENT STRUCTURE ANALYSIS")
        content_parts.append("="*50)
        content_parts.append(f"File: {extracted_data['filename']}")
        content_parts.append(f"Sheets: {', '.join(extracted_data['metadata']['sheet_names'])}")
        content_parts.append("")
        
        # Add detailed sheet analysis
        for sheet_name, sheet_data in extracted_data['sheets'].items():
            content_parts.append(f"SHEET: {sheet_name}")
            content_parts.append("-" * 30)
            content_parts.append(f"Dimensions: {sheet_data['dimensions']['rows']} rows × {sheet_data['dimensions']['cols']} columns")
            content_parts.append(f"Merged cells: {len(sheet_data['merged_cells'])}")
            content_parts.append("")
            
            # Show merged cells structure first
            if sheet_data['merged_cells']:
                content_parts.append("MERGED CELLS:")
                for merged in sheet_data['merged_cells']:
                    rows_span = merged['end_row'] - merged['start_row'] + 1
                    cols_span = merged['end_col'] - merged['start_col'] + 1
                    content_parts.append(f"  • {merged['range']} ({rows_span}×{cols_span}): {merged['value']}")
                content_parts.append("")
            
            # Show content in a more readable format
            content_parts.append("CONTENT STRUCTURE:")
            grid = sheet_data['grid']
            
            # Group by logical sections based on merged cells and content
            current_section = []
            for row_idx, row in enumerate(grid):
                if any(cell.strip() for cell in row if cell):
                    row_content = []
                    for col_idx, cell in enumerate(row):
                        if cell.strip():
                            row_content.append(f"[{row_idx+1},{col_idx+1}] {cell}")
                    
                    if row_content:
                        current_section.append("  " + " | ".join(row_content))
                        
                        # Add section break after significant gaps
                        if len(current_section) > 0 and self._is_section_break(grid, row_idx):
                            content_parts.extend(current_section)
                            content_parts.append("")
                            current_section = []
            
            # Add remaining content
            if current_section:
                content_parts.extend(current_section)
            
            content_parts.append("\n" + "="*60 + "\n")
        
        return "\n".join(content_parts)
    
    def _is_section_break(self, grid: List[List[str]], current_row: int) -> bool:
        """
        Determine if there should be a section break after current row
        """
        # Check if next few rows are empty (indicates section break)
        for i in range(1, min(3, len(grid) - current_row)):
            next_row = grid[current_row + i]
            if any(cell.strip() for cell in next_row if cell):
                return False
        return True
    
    def is_supported_file(self, file_path: str) -> bool:
        """Check if file format is supported"""
        ext = os.path.splitext(file_path)[1].lower()
        return ext in self.supported_extensions
    
    def get_supported_extensions(self) -> List[str]:
        """Get list of supported file extensions"""
        return self.supported_extensions.copy() 
    
    def to_hybrid_markdown(self, extracted_data: Dict[str, Any]) -> str:
        """
        Convert extracted data to hybrid markdown format:
        - Plain text for merged header cells
        - Markdown tables for data sections
        """
        content_parts = []
        
        # Add header information
        content_parts.append(f"# {extracted_data['filename']}")
        content_parts.append("")
        
        # Process each sheet
        for sheet_name, sheet_data in extracted_data['sheets'].items():
            if len(extracted_data['sheets']) > 1:
                content_parts.append(f"## Sheet: {sheet_name}")
                content_parts.append("")
            
            grid = sheet_data['grid']
            merged_cells = sheet_data['merged_cells']
            
            # Identify header sections and data table sections
            header_rows, data_sections = self._identify_sections(grid, merged_cells)
            
            # Process each section
            current_row = 0
            
            for section_type, start_row, end_row in data_sections:
                # Add any header rows before this section
                while current_row < start_row:
                    if current_row in header_rows:
                        header_content = self._format_header_row(grid[current_row], current_row, merged_cells)
                        if header_content:
                            content_parts.append(header_content)
                            content_parts.append("")
                    current_row += 1
                
                # Process the section
                if section_type == 'table':
                    table_content = self._format_table_section(grid, start_row, end_row, merged_cells)
                    if table_content:
                        content_parts.extend(table_content)
                        content_parts.append("")
                elif section_type == 'header':
                    for row_idx in range(start_row, end_row + 1):
                        header_content = self._format_header_row(grid[row_idx], row_idx, merged_cells)
                        if header_content:
                            content_parts.append(header_content)
                    content_parts.append("")
                
                current_row = end_row + 1
            
            # Process remaining header rows
            while current_row < len(grid):
                if current_row in header_rows:
                    header_content = self._format_header_row(grid[current_row], current_row, merged_cells)
                    if header_content:
                        content_parts.append(header_content)
                        content_parts.append("")
                current_row += 1
        
        return "\n".join(content_parts)
    
    def _identify_sections(self, grid: List[List[str]], merged_cells: List[Dict]) -> Tuple[set, List[Tuple[str, int, int]]]:
        """
        Identify header rows and data table sections
        Returns: (header_rows_set, sections_list)
        """
        header_rows = set()
        sections = []
        
        # Find rows with large merged cells (likely headers)
        for merged in merged_cells:
            start_row = merged['start_row'] - 1  # Convert to 0-based
            end_row = merged['end_row'] - 1
            cols_span = merged['end_col'] - merged['start_col'] + 1
            
            # If merged cell spans many columns (>= 3), consider it a header
            if cols_span >= 3:
                for row in range(start_row, end_row + 1):
                    header_rows.add(row)
        
        # Find rows that are mostly empty or have very few cells
        for row_idx, row in enumerate(grid):
            non_empty_cells = sum(1 for cell in row if cell.strip())
            if non_empty_cells <= 2 and non_empty_cells > 0:
                header_rows.add(row_idx)
        
        # Identify table sections (consecutive rows with many columns)
        current_section_start = None
        current_section_type = None
        
        for row_idx, row in enumerate(grid):
            non_empty_cells = sum(1 for cell in row if cell.strip())
            
            if row_idx in header_rows:
                # This is a header row
                if current_section_type == 'table' and current_section_start is not None:
                    sections.append(('table', current_section_start, row_idx - 1))
                    current_section_start = None
                
                if current_section_type != 'header':
                    if current_section_start is not None:
                        sections.append((current_section_type, current_section_start, row_idx - 1))
                    current_section_start = row_idx
                    current_section_type = 'header'
            
            elif non_empty_cells >= 3:  # Likely a data row
                if current_section_type != 'table':
                    if current_section_start is not None:
                        sections.append((current_section_type, current_section_start, row_idx - 1))
                    current_section_start = row_idx
                    current_section_type = 'table'
            
            else:
                # Empty or sparse row
                if current_section_start is not None:
                    sections.append((current_section_type, current_section_start, row_idx - 1))
                    current_section_start = None
                    current_section_type = None
        
        # Close final section
        if current_section_start is not None:
            sections.append((current_section_type, current_section_start, len(grid) - 1))
        
        return header_rows, sections
    
    def _format_header_row(self, row: List[str], row_idx: int, merged_cells: List[Dict]) -> str:
        """
        Format a header row as plain text
        """
        # Find merged cells in this row
        row_merged_cells = []
        for merged in merged_cells:
            if merged['start_row'] - 1 <= row_idx <= merged['end_row'] - 1:
                row_merged_cells.append(merged)
        
        # If there are large merged cells, use their values
        for merged in row_merged_cells:
            cols_span = merged['end_col'] - merged['start_col'] + 1
            if cols_span >= 3:  # Large merged cell
                text = merged['value'].strip()
                if text:
                    # Determine header level based on span
                    if cols_span >= 10:
                        return f"# {text}"
                    elif cols_span >= 6:
                        return f"## {text}"
                    else:
                        return f"### {text}"
        
        # Otherwise, concatenate non-empty cells
        non_empty_cells = [cell.strip() for cell in row if cell.strip()]
        if non_empty_cells:
            return f"**{' - '.join(non_empty_cells)}**"
        
        return ""
    
    def _format_table_section(self, grid: List[List[str]], start_row: int, end_row: int, merged_cells: List[Dict]) -> List[str]:
        """
        Format a table section as markdown table
        """
        if start_row > end_row:
            return []
        
        table_lines = []
        
        # Find the maximum number of non-empty columns in this section
        max_cols = 0
        for row_idx in range(start_row, end_row + 1):
            row = grid[row_idx]
            last_non_empty = -1
            for col_idx, cell in enumerate(row):
                if cell.strip():
                    last_non_empty = col_idx
            max_cols = max(max_cols, last_non_empty + 1)
        
        if max_cols == 0:
            return []
        
        # Create merged cell lookup for this section
        section_merged = {}
        for merged in merged_cells:
            for row in range(merged['start_row'] - 1, merged['end_row']):
                for col in range(merged['start_col'] - 1, merged['end_col']):
                    if start_row <= row <= end_row:
                        section_merged[(row, col)] = merged['value']
        
        # Generate table rows
        processed_rows = []
        for row_idx in range(start_row, end_row + 1):
            row = grid[row_idx]
            table_row = []
            
            for col_idx in range(max_cols):
                if (row_idx, col_idx) in section_merged:
                    # Use merged cell value
                    cell_value = section_merged[(row_idx, col_idx)]
                elif col_idx < len(row):
                    # Use regular cell value
                    cell_value = row[col_idx]
                else:
                    cell_value = ""
                
                # Clean up cell value
                cell_value = cell_value.strip().replace('|', '\\|')  # Escape pipes
                if not cell_value:
                    cell_value = " "
                
                table_row.append(cell_value)
            
            processed_rows.append(table_row)
        
        # Create markdown table
        if processed_rows:
            # Header row
            header_row = "| " + " | ".join(processed_rows[0]) + " |"
            table_lines.append(header_row)
            
            # Separator row
            separator = "| " + " | ".join(["---"] * max_cols) + " |"
            table_lines.append(separator)
            
            # Data rows
            for row in processed_rows[1:]:
                data_row = "| " + " | ".join(row) + " |"
                table_lines.append(data_row)
        
        return table_lines 
    
    def to_llm_optimized(self, extracted_data: Dict[str, Any]) -> str:
        """
        Convert to LLM-optimized format preserving Excel structure and order
        """
        content_parts = []
        
        # Process each sheet
        for sheet_name, sheet_data in extracted_data['sheets'].items():
            grid = sheet_data['grid']
            merged_cells = sheet_data['merged_cells']
            
            # Create a merged cells map for easy lookup by position
            merged_map = {}
            for merged in merged_cells:
                for row in range(merged['start_row'] - 1, merged['end_row']):
                    for col in range(merged['start_col'] - 1, merged['end_col']):
                        merged_map[(row, col)] = merged['value']
            
            # Process Excel from top to bottom, maintaining original structure
            current_section = []
            table_processed = False
            skip_until_row = -1  # Track which rows to skip after table processing
            
            row_idx = 0
            while row_idx < len(grid):
                # Skip rows that were already processed by table
                if row_idx <= skip_until_row:
                    row_idx += 1
                    continue
                    
                row = grid[row_idx]
                
                # Check if this is a merged cell row that spans many columns (header/title)
                is_title_row = False
                row_content = []
                
                for col_idx, cell in enumerate(row):
                    if (row_idx, col_idx) in merged_map:
                        # This cell is part of a merged area
                        merged_value = merged_map[(row_idx, col_idx)]
                        if merged_value and str(merged_value).strip():
                            # Find the size of this merged area
                            matching_merged = None
                            for merged in merged_cells:
                                if (merged['start_row'] - 1 <= row_idx < merged['end_row'] and 
                                    merged['start_col'] - 1 <= col_idx < merged['end_col']):
                                    matching_merged = merged
                                    break
                            
                            if matching_merged:
                                cols_span = matching_merged['end_col'] - matching_merged['start_col'] + 1
                                # If merged cell spans many columns, treat as title/header
                                if cols_span >= 6:
                                    is_title_row = True
                                    if str(merged_value).strip() not in [item.strip() for item in current_section]:
                                        current_section.append(str(merged_value).strip())
                                else:
                                    # Small merged cell, treat as regular data
                                    row_content.append(str(merged_value).strip())
                            break  # Skip other cells in this merged area
                    elif cell and str(cell).strip() and str(cell).strip() != 'None':
                        row_content.append(str(cell).strip())
                    else:
                        row_content.append("")
                
                # If it's not a title row, check if it's a table row
                if not is_title_row and row_content:
                    # Count non-empty cells to determine if this looks like a data table
                    non_empty_count = sum(1 for item in row_content if item.strip())
                    
                    if non_empty_count >= 5 and not table_processed:  # Likely a table row and haven't processed table yet
                        # This might be start of a table - check if we have accumulated section content
                        if current_section:
                            # Add section headers
                            for item in current_section:
                                content_parts.append(item)
                            content_parts.append("")  # Blank line after headers
                            current_section = []
                        
                        # Start processing table
                        table_content = self._extract_table_from_position(grid, merged_cells, row_idx)
                        if table_content:
                            content_parts.extend(table_content)
                            content_parts.append("")  # Blank line after table
                            table_processed = True
                            
                            # Calculate exactly which rows were consumed by table processing
                            table_end_row = self._find_table_end(grid, row_idx)
                            skip_until_row = table_end_row  # Skip all rows processed by table
                            
                    else:
                        # Regular content row - collect it
                        if non_empty_count >= 1:  # Any meaningful content
                            row_text = " ".join(item for item in row_content if item.strip())
                            if row_text.strip() and len(row_text.strip()) > 1:
                                current_section.append(row_text.strip())
                
                row_idx += 1
            
            # Add any remaining section content
            if current_section:
                # Check if we need a section header
                if table_processed and current_section:
                    content_parts.append("**Phạm vi báo cáo:**")
                    content_parts.append("")
                
                for item in current_section:
                    content_parts.append(item)
        
        return "\n".join(content_parts).strip()
    
    def _extract_table_from_position(self, grid: List[List[str]], merged_cells: List[Dict], start_row: int) -> List[str]:
        """
        Extract a well-formatted table starting from the given position
        """
        table_lines = []
        
        # Find the extent of this table
        table_end_row = start_row
        max_cols = 0
        
        # Determine table boundaries
        for row_idx in range(start_row, len(grid)):
            row = grid[row_idx]
            non_empty_count = sum(1 for cell in row if cell and str(cell).strip() and str(cell).strip() != 'None')
            
            if non_empty_count >= 3:  # Continue table
                table_end_row = row_idx
                max_cols = max(max_cols, len([cell for cell in row if cell and str(cell).strip()]))
            elif non_empty_count == 0:
                # Empty row might indicate end of table
                # Check next row to be sure
                if row_idx + 1 < len(grid):
                    next_row = grid[row_idx + 1]
                    next_non_empty = sum(1 for cell in next_row if cell and str(cell).strip() and str(cell).strip() != 'None')
                    if next_non_empty < 3:
                        break  # End of table
                else:
                    break
        
        if table_end_row <= start_row:
            return []
        
        # Extract table headers (first row) - ONLY non-empty columns
        header_row = grid[start_row]
        headers = []
        header_column_indices = []  # Track which original columns we keep
        
        for col_idx, cell in enumerate(header_row):
            if cell and str(cell).strip() and str(cell).strip() != 'None':
                headers.append(str(cell).strip())
                header_column_indices.append(col_idx)
        
        if not headers:
            return []
        
        # Create markdown table with ONLY meaningful columns
        table_lines.append("| " + " | ".join(headers) + " |")
        table_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
        
        # Add data rows - use SAME column indices as headers
        for row_idx in range(start_row + 1, table_end_row + 1):
            row = grid[row_idx]
            row_data = []
            
            # Extract data ONLY from columns that have headers
            for col_idx in header_column_indices:
                if col_idx < len(row) and row[col_idx] and str(row[col_idx]).strip() and str(row[col_idx]).strip() != 'None':
                    cell_content = str(row[col_idx]).strip()
                    # Handle multiline content in cells
                    cell_content = cell_content.replace('\n', '<br/>')
                    row_data.append(cell_content)
                else:
                    row_data.append("")
            
            # Only add row if it has some content
            if any(cell.strip() for cell in row_data):
                # Replace empty cells with "-" for better table formatting
                display_row = [cell if cell.strip() else "-" for cell in row_data]
                table_lines.append("| " + " | ".join(display_row) + " |")
        
        return table_lines
    
    def _find_table_end(self, grid: List[List[str]], start_row: int) -> int:
        """
        Find where the table ends
        """
        table_end_row = start_row
        
        for row_idx in range(start_row, len(grid)):
            row = grid[row_idx]
            non_empty_count = sum(1 for cell in row if cell and str(cell).strip() and str(cell).strip() != 'None')
            
            if non_empty_count >= 3:  # Continue table
                table_end_row = row_idx
            elif non_empty_count == 0:
                # Empty row might indicate end of table
                # Check next row to be sure
                if row_idx + 1 < len(grid):
                    next_row = grid[row_idx + 1]
                    next_non_empty = sum(1 for cell in next_row if cell and str(cell).strip() and str(cell).strip() != 'None')
                    if next_non_empty < 3:
                        break  # End of table
                else:
                    break
        
        return table_end_row
    
    def _extract_document_metadata(self, merged_cells: List[Dict]) -> Dict[str, str]:
        """Extract document metadata from large merged cells"""
        metadata = {}
        
        for merged in merged_cells:
            cols_span = merged['end_col'] - merged['start_col'] + 1
            value = merged['value'].strip()
            
            if cols_span >= 6 and value:  # Large merged cells likely contain metadata
                if 'ngân hàng' in value.lower():
                    metadata['Bank'] = value
                elif 'chi nhánh' in value.lower():
                    metadata['Branch'] = value.replace('Chi nhánh:', '').strip()
                elif 'phòng' in value.lower():
                    metadata['Department'] = value.replace('Phòng:', '').strip()
                elif 'ngày' in value.lower() and 'giờ' in value.lower():
                    metadata['Print Date'] = value.replace('Ngày giờ in:', '').strip()
                elif 'bảng kê' in value.lower() or 'báo cáo' in value.lower():
                    metadata['Report Title'] = value
        
        return metadata
    
    def _identify_clean_data_tables(self, grid: List[List[str]], merged_cells: List[Dict]) -> List[Dict]:
        """Identify actual data tables (not metadata)"""
        tables = []
        
        # Find rows with many non-empty columns (likely data)
        potential_table_rows = []
        for row_idx, row in enumerate(grid):
            non_empty = sum(1 for cell in row if cell.strip())
            if non_empty >= 5:  # Tables typically have many columns
                potential_table_rows.append(row_idx)
        
        if not potential_table_rows:
            return tables
        
        # Group consecutive rows into tables
        current_table_start = potential_table_rows[0]
        current_table_end = potential_table_rows[0]
        
        for i in range(1, len(potential_table_rows)):
            if potential_table_rows[i] - potential_table_rows[i-1] <= 2:  # Allow 1-2 empty rows
                current_table_end = potential_table_rows[i]
            else:
                # End current table, start new one
                if current_table_end - current_table_start >= 1:  # At least 2 rows
                    # Find title from merged cells above this table
                    title = self._find_table_title(current_table_start, merged_cells)
                    tables.append({
                        'title': title,
                        'start_row': current_table_start,
                        'end_row': current_table_end
                    })
                
                current_table_start = potential_table_rows[i]
                current_table_end = potential_table_rows[i]
        
        # Add final table
        if current_table_end - current_table_start >= 1:
            title = self._find_table_title(current_table_start, merged_cells)
            tables.append({
                'title': title,
                'start_row': current_table_start,
                'end_row': current_table_end
            })
        
        return tables
    
    def _find_table_title(self, table_start_row: int, merged_cells: List[Dict]) -> str:
        """Find title for a data table from nearby merged cells"""
        # Look for merged cells in the 5 rows above the table
        for merged in merged_cells:
            if (merged['start_row'] - 1 >= table_start_row - 5 and 
                merged['start_row'] - 1 < table_start_row):
                cols_span = merged['end_col'] - merged['start_col'] + 1
                if cols_span >= 3:  # Reasonably wide merged cell
                    value = merged['value'].strip()
                    if value and len(value) > 5:  # Meaningful text
                        return value
        
        return "Data Table"
    
    def _create_clean_table(self, grid: List[List[str]], start_row: int, end_row: int, merged_cells: List[Dict]) -> List[str]:
        """Create a clean table without duplicate columns"""
        if start_row > end_row:
            return []
        
        # Find actual columns by analyzing header patterns
        header_row_idx = start_row
        header_row = grid[header_row_idx]
        
        # Remove duplicate columns
        clean_columns = []
        seen_headers = set()
        
        for col_idx, cell in enumerate(header_row):
            if cell.strip():
                # Clean header name
                clean_header = cell.strip()
                # Remove duplicate markers like "_A", "_B" 
                base_header = clean_header.replace('_A', '').replace('_B', '').strip()
                
                if base_header not in seen_headers:
                    clean_columns.append({
                        'header': base_header,
                        'original_col': col_idx
                    })
                    seen_headers.add(base_header)
        
        if not clean_columns:
            return []
        
        table_lines = []
        
        # Create header row
        headers = [col['header'] for col in clean_columns]
        header_line = "| " + " | ".join(headers) + " |"
        table_lines.append(header_line)
        
        # Separator
        separator = "| " + " | ".join(["---"] * len(headers)) + " |"
        table_lines.append(separator)
        
        # Data rows
        for row_idx in range(start_row + 1, end_row + 1):
            if row_idx < len(grid):
                row = grid[row_idx]
                data_cells = []
                
                for col_info in clean_columns:
                    col_idx = col_info['original_col']
                    if col_idx < len(row):
                        cell_value = row[col_idx].strip()
                        # Clean up cell value
                        if not cell_value or cell_value.lower() == 'none':
                            cell_value = "-"
                        else:
                            # Remove line breaks and clean
                            cell_value = cell_value.replace('\n', ' ').strip()
                            # Escape pipes
                            cell_value = cell_value.replace('|', '\\|')
                        data_cells.append(cell_value)
                    else:
                        data_cells.append("-")
                
                # Only add row if it has meaningful data
                if any(cell != "-" for cell in data_cells):
                    data_line = "| " + " | ".join(data_cells) + " |"
                    table_lines.append(data_line)
        
        return table_lines
    
    def _extract_business_rules(self, grid: List[List[str]], merged_cells: List[Dict]) -> List[str]:
        """Extract business rules and formulas"""
        rules = []
        
        # Look for cells containing logical expressions or formulas
        for row in grid:
            for cell in row:
                if cell and cell.strip():
                    cell_content = cell.strip()
                    # Look for conditional logic
                    if ('IF ' in cell_content.upper() or 
                        'THEN' in cell_content.upper() or
                        'ELSE' in cell_content.upper() or
                        '=' in cell_content and 'LNP' in cell_content.upper()):
                        if cell_content not in rules:
                            rules.append(cell_content)
                    # Look for UNION operations
                    elif 'UNION' in cell_content.upper():
                        if cell_content not in rules:
                            rules.append(cell_content)
        
        return rules
    
    def _extract_scope_information(self, grid: List[List[str]]) -> List[str]:
        """Extract report scope and context information"""
        scope_info = []
        
        for row in grid:
            for cell in row:
                if cell and cell.strip():
                    cell_content = cell.strip()
                    # Look for scope definitions
                    if ('phạm vi' in cell_content.lower() or
                        'bản ghi' in cell_content.lower() or
                        'key:' in cell_content.lower()):
                        if cell_content not in scope_info:
                            scope_info.append(cell_content)
        
        return scope_info 