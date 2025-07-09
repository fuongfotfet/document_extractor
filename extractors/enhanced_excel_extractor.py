"""
Enhanced Excel Extractor - Better handling of merged cells
"""

import os
import pandas as pd
from openpyxl import load_workbook
from typing import Dict, List, Any, Tuple
from .base_extractor import BaseExtractor


class EnhancedExcelExtractor(BaseExtractor):
    """
    Enhanced Excel extractor with better merged cell handling
    """
    
    def __init__(self):
        super().__init__()
        self.supported_extensions = ['.xlsx', '.xls']
    
    def extract(self, file_path: str) -> Dict[str, Any]:
        """
        Extract data from Excel file with better merged cell handling
        """
        self.validate_file_exists(file_path)
        
        if not self.is_supported_file(file_path):
            raise ValueError(f"File extension not supported. Expected: {self.supported_extensions}")
        
        result = {
            'filename': os.path.basename(file_path),
            'file_type': 'excel',
            'sheets': [],
            'metadata': {}
        }
        
        try:
            if file_path.endswith('.xlsx'):
                # Use openpyxl for better merged cell handling
                workbook = load_workbook(file_path, data_only=True)
                result['metadata']['total_sheets'] = len(workbook.sheetnames)
                result['metadata']['sheet_names'] = workbook.sheetnames
                
                for sheet_name in workbook.sheetnames:
                    sheet_data = self._process_sheet_with_merged_cells(workbook[sheet_name])
                    sheet_data['sheet_name'] = sheet_name
                    result['sheets'].append(sheet_data)
            else:
                # Fallback to pandas for .xls files
                excel_file = pd.ExcelFile(file_path, engine='xlrd')
                result['metadata']['total_sheets'] = len(excel_file.sheet_names)
                result['metadata']['sheet_names'] = excel_file.sheet_names
                
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                    sheet_data = {
                        'sheet_name': sheet_name,
                        'shape': df.shape,
                        'cleaned_data': self._clean_duplicate_values(df.values.tolist()),
                        'merged_regions': []  # .xls doesn't provide merged cell info easily
                    }
                    result['sheets'].append(sheet_data)
                    
        except Exception as e:
            raise Exception(f"Error processing Excel file: {str(e)}")
        
        return result
    
    def _process_sheet_with_merged_cells(self, worksheet) -> Dict[str, Any]:
        """
        Process worksheet with proper merged cell handling
        """
        # Get worksheet dimensions
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Create data matrix
        data_matrix = []
        for row in range(1, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                row_data.append(cell.value)
            data_matrix.append(row_data)
        
        # Get merged cell ranges
        merged_ranges = list(worksheet.merged_cells.ranges)
        
        # Clean the data matrix by removing duplicates from merged cells
        cleaned_data = self._clean_merged_cell_duplicates(data_matrix, merged_ranges)
        
        return {
            'shape': (max_row, max_col),
            'cleaned_data': cleaned_data,
            'merged_regions': [str(range) for range in merged_ranges],
            'original_data': data_matrix
        }
    
    def _clean_merged_cell_duplicates(self, data_matrix: List[List], merged_ranges) -> List[List]:
        """
        Clean duplicate values from merged cells
        """
        cleaned_matrix = [row[:] for row in data_matrix]  # Deep copy
        
        for merged_range in merged_ranges:
            # Get the top-left cell value (the original value)
            min_row, min_col = merged_range.min_row - 1, merged_range.min_col - 1
            max_row, max_col = merged_range.max_row - 1, merged_range.max_col - 1
            
            if min_row < len(cleaned_matrix) and min_col < len(cleaned_matrix[0]):
                original_value = cleaned_matrix[min_row][min_col]
                
                # Clear duplicates in the merged range, keep only top-left
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        if row < len(cleaned_matrix) and col < len(cleaned_matrix[row]):
                            if row == min_row and col == min_col:
                                # Keep the original value in top-left
                                continue
                            else:
                                # Clear duplicates
                                cleaned_matrix[row][col] = None
        
        return cleaned_matrix
    
    def _clean_duplicate_values(self, data_list: List[List]) -> List[List]:
        """
        Simple duplicate cleaning for when we can't detect merged cells
        """
        if not data_list:
            return data_list
        
        cleaned_data = []
        for row in data_list:
            cleaned_row = []
            prev_value = None
            consecutive_count = 0
            
            for cell in row:
                if cell == prev_value and cell is not None and str(cell).strip():
                    consecutive_count += 1
                    # If we see the same value 3+ times consecutively, likely a merged cell
                    if consecutive_count >= 3:
                        cleaned_row.append(None)  # Replace with None
                    else:
                        cleaned_row.append(cell)
                else:
                    consecutive_count = 1
                    cleaned_row.append(cell)
                prev_value = cell
            
            cleaned_data.append(cleaned_row)
        
        return cleaned_data
    
    def to_markdown(self, extracted_data: Dict[str, Any]) -> str:
        """
        Convert to markdown with better table formatting
        """
        markdown_content = []
        
        for sheet in extracted_data['sheets']:
            if len(extracted_data['sheets']) > 1:
                markdown_content.append(f"## {sheet['sheet_name']}")
                markdown_content.append("")
            
            # Add merged cell info if available
            if sheet.get('merged_regions'):
                markdown_content.append("**Merged Cell Regions:**")
                for region in sheet['merged_regions'][:5]:  # Show first 5
                    markdown_content.append(f"- {region}")
                markdown_content.append("")
            
            # Convert cleaned data to markdown table
            cleaned_data = sheet['cleaned_data']
            if cleaned_data:
                markdown_table = self._create_clean_markdown_table(cleaned_data)
                markdown_content.append(markdown_table)
                markdown_content.append("")
        
        return "\n".join(markdown_content).strip()
    
    def _create_clean_markdown_table(self, data: List[List]) -> str:
        """
        Create markdown table from cleaned data
        """
        if not data:
            return ""
        
        # Remove completely empty rows
        non_empty_data = [row for row in data if any(cell is not None and str(cell).strip() for cell in row)]
        
        if not non_empty_data:
            return ""
        
        # Find the maximum number of columns
        max_cols = max(len(row) for row in non_empty_data)
        
        # Normalize rows to same length
        normalized_data = []
        for row in non_empty_data:
            normalized_row = row + [None] * (max_cols - len(row))
            # Convert to strings and handle None values
            str_row = [str(cell) if cell is not None else "" for cell in normalized_row]
            normalized_data.append(str_row)
        
        if not normalized_data:
            return ""
        
        # Create markdown table
        table_lines = []
        
        # Add header row (first row)
        header = "| " + " | ".join(normalized_data[0]) + " |"
        table_lines.append(header)
        
        # Add separator
        separator = "|" + "|".join(["-" * max(1, len(cell)) for cell in normalized_data[0]]) + "|"
        table_lines.append(separator)
        
        # Add data rows
        for row in normalized_data[1:]:
            data_row = "| " + " | ".join(row) + " |"
            table_lines.append(data_row)
        
        return "\n".join(table_lines) 